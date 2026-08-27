"""
SSIM差异图 + 框检测 + 水位线提取 批量处理
输入：文件夹中的所有图片（相邻图片两两对比）
输出：每个图片对一个小文件夹 + Excel水位线报告
"""

import cv2
import numpy as np
import os
import glob
from PIL import Image
from skimage.metrics import structural_similarity as ssim
from ssim_box_detector import SSIMBoxDetector


def imread_unicode(path):
    """支持中文路径的图像读取"""
    try:
        img = Image.open(path)
        return np.array(img)
    except Exception as e:
        raise ValueError(f"无法读取图片: {path}, 错误: {e}")


def extract_water_level_from_mask(mask):
    """从掩码提取水位线（岸体下边缘）"""
    if mask.dtype != np.uint8:
        mask = mask.astype(np.uint8)

    h, w = mask.shape
    bank_line = []

    for x in range(w):
        col = mask[:, x]
        white_pixels = np.where(col > 0)[0]
        if len(white_pixels) > 0:
            bank_line.append(white_pixels[-1])
        else:
            bank_line.append(h)

    bank_line_array = np.array(bank_line)
    valid_lines = bank_line_array[bank_line_array < h]

    if len(valid_lines) > 0:
        water_level_y = float(np.median(valid_lines))
    else:
        water_level_y = h // 2

    return bank_line, water_level_y


def process_image_pair(img_path1, img_path2, output_dir):
    """处理一对图片"""
    name1 = os.path.splitext(os.path.basename(img_path1))[0]
    name2 = os.path.splitext(os.path.basename(img_path2))[0]
    pair_name = f"{name1}_{name2}"

    pair_dir = os.path.join(output_dir, pair_name)
    os.makedirs(pair_dir, exist_ok=True)

    result = {
        'pair': pair_name,
        'img1': os.path.basename(img_path1),
        'img2': os.path.basename(img_path2),
        'success': False,
        'water_level_y': None,
        'ssim_diff_path': None,
        'box_img_path': None
    }

    try:
        # SSIM对比
        img1 = imread_unicode(img_path1)
        img2 = imread_unicode(img_path2)

        if len(img1.shape) == 3:
            img1 = cv2.cvtColor(img1, cv2.COLOR_RGB2GRAY)
        if len(img2.shape) == 3:
            img2 = cv2.cvtColor(img2, cv2.COLOR_RGB2GRAY)

        if img1.shape != img2.shape:
            img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]))

        ssim_score, ssim_map = ssim(img1, img2, full=True, win_size=11, data_range=255)
        diff_map = 1 - ssim_map
        diff_norm = (diff_map * 255).astype(np.uint8)
        _, bool_matrix = cv2.threshold(diff_norm, thresh=0, maxval=255, type=cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        auto_threshold = _ / 255.0

        ssim_diff_path = os.path.join(pair_dir, "ssim_diff.png")
        cv2.imwrite(ssim_diff_path, bool_matrix)
        result['ssim_diff_path'] = ssim_diff_path
        result['threshold'] = auto_threshold

        # 框检测 - 获取水位线（黑白分界线Y值）
        detector = SSIMBoxDetector()
        detector.load_from_array(bool_matrix)
        boxes = detector.detect_two_boxes()

        # 水位线 = POS框的Y2 = NEG框的Y1 = 黑白分界线
        water_level_y = None
        box_info = ""
        for box in boxes:
            x1, y1, x2, y2, label = box
            if label == True:  # POS = 岸体
                water_level_y = y2  # 岸体下边缘 = 水位线
                box_info = f"POS:({x1},{y1},{x2},{y2}) NEG:({boxes[1][0]},{boxes[1][1]},{boxes[1][2]},{boxes[1][3]})"
            if label == False:  # NEG = 水面
                water_level_y = y1  # 水面起始 = 水位线

        box_img_path = os.path.join(pair_dir, "box_detection.png")
        detector.visualize(save_path=box_img_path)
        result['box_img_path'] = box_img_path
        result['boxes'] = boxes
        result['water_level_y'] = water_level_y
        result['box_info'] = box_info
        result['success'] = True

        print(f"  ✓ {pair_name}: 水位Y={water_level_y}, 阈值={auto_threshold:.4f}")

    except Exception as e:
        result['error'] = str(e)
        print(f"  ✗ {pair_name}: {e}")

    return result


def save_to_excel(results, excel_path):
    """保存结果到Excel（追加模式，不覆盖原数据）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        print(f"追加到已存在的Excel: {excel_path}")
        start_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "水位检测结果"
        print(f"创建新Excel: {excel_path}")
        start_row = 2

    # 查找或创建新列
    water_col = None
    threshold_col = None
    box_col = None
    for col in range(1, ws.max_column + 20):
        val = ws.cell(row=1, column=col).value
        if val == "水位Y":
            water_col = col
        if val == "自适应阈值":
            threshold_col = col
        if val == "框坐标":
            box_col = col

    if water_col is None:
        water_col = ws.max_column + 1
        ws.cell(row=1, column=water_col, value="水位Y")
        ws.cell(row=1, column=water_col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=1, column=water_col).font = Font(bold=True, color="FFFFFF")

    if threshold_col is None:
        threshold_col = ws.max_column + 1
        ws.cell(row=1, column=threshold_col, value="自适应阈值")
        ws.cell(row=1, column=threshold_col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=1, column=threshold_col).font = Font(bold=True, color="FFFFFF")

    if box_col is None:
        box_col = ws.max_column + 1
        ws.cell(row=1, column=box_col, value="框坐标")
        ws.cell(row=1, column=box_col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=1, column=box_col).font = Font(bold=True, color="FFFFFF")

    # 填充数据（从最后一行开始追加）
    for i, result in enumerate(results):
        row_idx = start_row + i
        water_y = str(result['water_level_y']) if result['water_level_y'] is not None else ""
        threshold = f"{result.get('threshold', ''):.4f}" if result.get('threshold') else ""
        box_info = result.get('box_info', "")

        ws.cell(row=row_idx, column=water_col, value=water_y)
        ws.cell(row=row_idx, column=threshold_col, value=threshold)
        ws.cell(row=row_idx, column=box_col, value=box_info)

    wb.save(excel_path)
    print(f"已保存 {len(results)} 条数据到: {excel_path}")


def batch_process_folder(input_dir, output_dir, excel_path):
    """批量处理文件夹中的图片"""
    os.makedirs(output_dir, exist_ok=True)

    extensions = ['*.png', '*.jpg', '*.jpeg', '*.bmp', '*.tiff', '*.tif']
    image_files = []
    for ext in extensions:
        image_files.extend(glob.glob(os.path.join(input_dir, ext)))
        image_files.extend(glob.glob(os.path.join(input_dir, ext.upper())))

    def extract_number(path):
        import re
        nums = re.findall(r'\d+', os.path.basename(path))
        return int(nums[0]) if nums else 0

    image_files = sorted(list(set(image_files)), key=extract_number)
    print(f"找到 {len(image_files)} 张图片")
    print(f"输入: {input_dir}")
    print(f"输出: {output_dir}")
    print("=" * 50)

    if len(image_files) < 2:
        print("错误: 需要至少2张图片")
        return

    results = []

    for i in range(len(image_files) - 1):
        img1 = image_files[i]
        img2 = image_files[i + 1]
        print(f"\n[{i+1}/{len(image_files)-1}] {os.path.basename(img1)} <-> {os.path.basename(img2)}")

        result = process_image_pair(img1, img2, output_dir)
        results.append(result)

    # 保存Excel
    save_to_excel(results, excel_path)

    success_count = sum(1 for r in results if r['success'])
    print("\n" + "=" * 50)
    print("处理完成!")
    print(f"成功: {success_count}/{len(results)}")
    print(f"输出目录: {output_dir}")

    return results


if __name__ == "__main__":
    INPUT_DIR = r"E:\VS code\SAM3\ROI_input"
    OUTPUT_DIR = r"E:\VS code\SAM3\flood_mask_output"
    EXCEL_PATH = r"E:\VS code\SAM3\flood_water_level.xlsx"

    batch_process_folder(INPUT_DIR, OUTPUT_DIR, EXCEL_PATH)
